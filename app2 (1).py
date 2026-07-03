import streamlit as st
import google.generativeai as genai
import json
import re
import io
import docx
import logging
from datetime import datetime
import openpyxl
import pandas as pd
import unicodedata

# ============================================================
# 1. CONFIGURAÇÃO DE LOGS
# ============================================================
logging.basicConfig(
    filename="automacao_rh_raquel.log",
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
    encoding="utf-8"
)

st.set_page_config(page_title="Automação de Relatórios RH", page_icon="📊", layout="wide")
st.title("📊 Sistema de Automação de Relatórios — RH")
st.subheader("Extração de Notas para Excel Modelo Predefinido")

# ============================================================
# 2. INICIALIZAÇÃO DA API DO GEMINI
# ============================================================
try:
    genai.configure(api_key=st.secrets["GOOGLE_API_KEY"])
except Exception:
    st.error("Erro: Chave API em falta nos Secrets do Streamlit. Configura 'GOOGLE_API_KEY' em Settings > Secrets.")
    st.stop()

# Modelo atual e estável (gemini-pro foi descontinuado pela Google em 2025/2026
# e devolve erro 404). gemini-2.5-flash é rápido, económico e adequado para
# extração de dados estruturados.
MODELO_GEMINI = "gemini-2.5-flash"

# ============================================================
# 3. ESTADO DA SESSÃO (evita perder o ficheiro gerado ao clicar em Download)
# ============================================================
if "dados_extraidos" not in st.session_state:
    st.session_state.dados_extraidos = None
if "excel_pronto" not in st.session_state:
    st.session_state.excel_pronto = None
if "nome_ficheiro_saida" not in st.session_state:
    st.session_state.nome_ficheiro_saida = None
if "linhas_inseridas" not in st.session_state:
    st.session_state.linhas_inseridas = 0


# ============================================================
# 4. LEITURA DE TEXTOS (Word e TXT)
# ============================================================
def ler_ficheiro_txt(uploaded_file):
    return uploaded_file.read().decode("utf-8", errors="ignore")


def ler_ficheiro_docx(uploaded_file):
    doc = docx.Document(io.BytesIO(uploaded_file.read()))
    return "\n".join([para.text for para in doc.paragraphs])


# ============================================================
# 5. AGENTE DE INTELIGÊNCIA ARTIFICIAL
# ============================================================
def extrair_dados_com_gemini(texto_notas, mapeamento_campos):
    model = genai.GenerativeModel(model_name=MODELO_GEMINI)

    prompt_sistema = f"""
    Tu és um Assistente de RH Avançado especialista em extração de dados estruturados.
    O teu objetivo é ler as notas fornecidas e extrair os dados necessários para preencher uma tabela corporativa.

    Deves mapear as informações estritamente para as seguintes colunas fornecidas pelo utilizador:
    {mapeamento_campos}

    Regras estritas de resposta:
    1. Responde APENAS com um array JSON válido contendo objetos. Cada objeto representa uma linha a introduzir no Excel.
    2. Usa exatamente o nome dos campos fornecidos como chaves do JSON (ex: "Nome", "Data", "Ocorrência", "Detalhes").
    3. Se a informação não existir nas notas para alguma das chaves, deixa o valor vazio "".
    4. NÃO adiciones nenhuma introdução, explicação ou texto fora do array JSON.
    5. Se identificares várias pessoas/ocorrências distintas nas notas, cria um objeto por cada uma.
    """

    try:
        response = model.generate_content(
            [prompt_sistema, f"Notas para analisar:\n\n{texto_notas}"],
            generation_config={"temperature": 0.1},
        )

        # Verifica se a resposta foi bloqueada ou está vazia antes de aceder a .text
        if not response.candidates:
            st.sidebar.error("O Gemini não devolveu nenhuma resposta (possível bloqueio de segurança).")
            logging.error("Resposta sem candidates. feedback=%s", getattr(response, "prompt_feedback", None))
            return None

        texto_resposta = (response.text or "").strip()

        if not texto_resposta:
            st.sidebar.error("O Gemini devolveu uma resposta vazia.")
            return None

        # Remove blocos de código Markdown (```json ... ```)
        if "```" in texto_resposta:
            texto_resposta = re.sub(r"```json|```", "", texto_resposta).strip()

        # Isola o array JSON caso o modelo tenha adicionado texto antes/depois
        match = re.search(r"\[.*\]", texto_resposta, re.DOTALL)
        if match:
            texto_resposta = match.group(0)

        dados = json.loads(texto_resposta)

        if not isinstance(dados, list):
            st.sidebar.error("A resposta da IA não é uma lista de registos. Tenta reformular as notas.")
            return None

        return dados

    except json.JSONDecodeError as e:
        logging.error(f"Erro ao interpretar JSON da IA: {e}\nResposta bruta: {texto_resposta}")
        st.sidebar.error(f"A IA devolveu um formato inválido. Detalhe: {e}")
        return None
    except Exception as e:
        logging.error(f"Erro na extração do Agente: {e}")
        st.sidebar.error(f"Detalhe do erro do Gemini: {e}")
        return None


# ============================================================
# 6. LOCALIZAÇÃO DO CABEÇALHO E ESCRITA NO EXCEL
# ============================================================
def normalizar(texto):
    """Remove acentos, espaços extra e diferenças de maiúsculas/minúsculas
    para permitir comparações fiáveis entre nomes de colunas."""
    if texto is None:
        return ""
    texto = str(texto).strip().lower()
    texto = unicodedata.normalize("NFKD", texto).encode("ascii", "ignore").decode("ascii")
    texto = re.sub(r"\s+", " ", texto)
    return texto


def encontrar_cabecalho(ws, colunas_alvo, limite_linhas=50):
    """Procura a linha de cabeçalho e devolve (linha, mapa_nome_alvo->coluna).
    A comparação é feita de forma normalizada (ignora acentos/maiúsculas/espaços),
    mas o mapa devolvido usa sempre o nome EXATO fornecido pelo utilizador como chave."""
    colunas_alvo_norm = {normalizar(c): c for c in colunas_alvo}

    for r in range(1, min(limite_linhas, ws.max_row) + 1):
        valores_linha_norm = []
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            valores_linha_norm.append(normalizar(cell.value))

        if any(cn in valores_linha_norm for cn in colunas_alvo_norm.keys()):
            mapa = {}
            for c in range(1, ws.max_column + 1):
                cell = ws.cell(row=r, column=c)
                nome_celula_norm = normalizar(cell.value)
                if nome_celula_norm in colunas_alvo_norm:
                    nome_original_utilizador = colunas_alvo_norm[nome_celula_norm]
                    mapa[nome_original_utilizador] = c
            return r, mapa

    return None, {}


def preencher_excel(excel_bytes, lista_dados, colunas_alvo):
    wb = openpyxl.load_workbook(io.BytesIO(excel_bytes))
    ws = wb.active

    linha_cabecalho, mapa_colunas_index = encontrar_cabecalho(ws, colunas_alvo)

    if not mapa_colunas_index:
        # Não arriscamos escrever "às cegas" por ordem de coluna, pois isso é o que
        # estava a fazer os dados aparecerem em colunas erradas / no fundo do ficheiro.
        # É preferível parar e avisar claramente do que corromper o Excel da empresa.
        raise ValueError(
            "Não foi possível encontrar nenhuma das colunas indicadas no cabeçalho do Excel. "
            "Verifica se os nomes em 'Colunas do Excel original' correspondem aos nomes reais "
            "das colunas no ficheiro (revê o painel de diagnóstico acima)."
        )

    if len(mapa_colunas_index) < len(colunas_alvo):
        colunas_nao_encontradas = [c for c in colunas_alvo if c not in mapa_colunas_index]
        st.warning(
            f"⚠️ Estas colunas não foram encontradas no cabeçalho e ficarão vazias: "
            f"{', '.join(colunas_nao_encontradas)}"
        )

    # Descobre a próxima linha vazia, saltando células fundidas
    proxima_linha = linha_cabecalho + 1
    while True:
        linha_vazia = True
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=proxima_linha, column=c)
            if type(cell).__name__ == "MergedCell" or cell.value is not None:
                linha_vazia = False
                break
        if linha_vazia:
            break
        proxima_linha += 1

    linhas_inseridas = 0
    for registo in lista_dados:
        # Normaliza as chaves do registo (a IA ou a edição manual podem devolver
        # nomes ligeiramente diferentes em maiúsculas/acentos/espaços)
        registo_norm = {normalizar(k): v for k, v in registo.items()}

        # Salta linhas cuja interseção caia numa MergedCell secundária
        tentativas = 0
        while any(
            type(ws.cell(row=proxima_linha, column=num_col)).__name__ == "MergedCell"
            for num_col in mapa_colunas_index.values()
        ):
            proxima_linha += 1
            tentativas += 1
            if tentativas > 1000:  # proteção contra loop infinito
                raise RuntimeError("Não foi possível encontrar linhas livres no Excel (demasiadas células fundidas).")

        for nome_campo, num_coluna in mapa_colunas_index.items():
            valor_final = registo_norm.get(normalizar(nome_campo), "")
            if valor_final is None or (isinstance(valor_final, float) and pd.isna(valor_final)):
                valor_final = ""
            cell = ws.cell(row=proxima_linha, column=num_coluna)
            if type(cell).__name__ != "MergedCell":
                cell.value = valor_final

        proxima_linha += 1
        linhas_inseridas += 1

    output_final = io.BytesIO()
    wb.save(output_final)
    output_final.seek(0)
    return output_final, linhas_inseridas


# ============================================================
# 7. INTERFACE DO UTILIZADOR
# ============================================================
col1, col2 = st.columns([1, 1])

with col1:
    st.markdown("### 📄 1. Documentos Base (Word/Notepad)")
    ficheiros_carregados = st.file_uploader(
        "Carrega as notas de RH", type=["txt", "docx"], accept_multiple_files=True
    )

    texto_acumulado = ""
    if ficheiros_carregados:
        for f in ficheiros_carregados:
            if f.name.endswith(".txt"):
                texto_acumulado += ler_ficheiro_txt(f) + "\n"
            elif f.name.endswith(".docx"):
                texto_acumulado += ler_ficheiro_docx(f) + "\n"

    texto_colado = st.text_area("Ou escreve/cola texto adicional aqui:", height=150)
    texto_final = (texto_acumulado + "\n" + texto_colado).strip()

with col2:
    st.markdown("### 📂 2. Configuração do Modelo da Empresa")
    excel_modelo = st.file_uploader("Carrega aqui o Excel Modelo/Template da Empresa (.xlsx)", type=["xlsx"])

    st.caption(
        "Identificação das Colunas (escreve os nomes exatamente como aparecem no cabeçalho "
        "do Excel original, separados por vírgulas):"
    )
    campos_predefinidos = "Nome, Data, Ocorrência, Detalhes"
    campos_usuario = st.text_input("Colunas do Excel original:", value=campos_predefinidos)

    # Diagnóstico: mostra logo se as colunas indicadas foram encontradas no Excel,
    # ANTES de gastar uma chamada à IA ou de escrever fosse o que fosse.
    if excel_modelo and campos_usuario.strip():
        try:
            excel_modelo.seek(0)
            wb_check = openpyxl.load_workbook(io.BytesIO(excel_modelo.read()), data_only=True)
            ws_check = wb_check.active
            colunas_alvo_check = [c.strip() for c in campos_usuario.split(",") if c.strip()]
            linha_check, mapa_check = encontrar_cabecalho(ws_check, colunas_alvo_check)

            if mapa_check and len(mapa_check) == len(colunas_alvo_check):
                st.success(f"✅ Cabeçalho encontrado na linha {linha_check}. Todas as colunas foram reconhecidas.")
                with st.expander("Ver mapeamento detetado"):
                    for nome, col in mapa_check.items():
                        st.write(f"**{nome}** → coluna {openpyxl.utils.get_column_letter(col)}")
            elif mapa_check:
                encontradas = list(mapa_check.keys())
                em_falta = [c for c in colunas_alvo_check if c not in encontradas]
                st.warning(
                    f"⚠️ Só foram reconhecidas {len(encontradas)} de {len(colunas_alvo_check)} colunas "
                    f"(linha {linha_check}). Em falta: {', '.join(em_falta)}. "
                    "Verifica se os nomes correspondem exatamente ao cabeçalho do Excel."
                )
            else:
                st.error(
                    "❌ Nenhuma das colunas indicadas foi encontrada no cabeçalho do Excel. "
                    "Confirma os nomes das colunas (revê o ficheiro carregado)."
                )
        except Exception as e:
            st.warning(f"Não foi possível pré-validar o Excel: {e}")

st.divider()

# ============================================================
# 8. PASSO 1 — ANALISAR NOTAS COM A IA
# ============================================================
if st.button("🤖 Analisar Notas", use_container_width=True, type="primary"):
    if not texto_final:
        st.warning("Introduz ou carrega notas primeiro.")
    elif not excel_modelo:
        st.error("Precisas de carregar o Excel Modelo da empresa para podermos preenchê-lo.")
    else:
        with st.spinner("O Agente está a ler as notas e a identificar os dados..."):
            lista_dados = extrair_dados_com_gemini(texto_final, campos_usuario)

        if lista_dados:
            st.session_state.dados_extraidos = lista_dados
            st.session_state.excel_pronto = None  # limpa resultado anterior
            st.success(f"✅ Foram identificados {len(lista_dados)} registo(s). Confirma os dados abaixo.")
        else:
            st.error(
                "Não foi possível extrair dados válidos. Verifica se o texto enviado contém dados claros "
                "ou consulta os logs."
            )

# ============================================================
# 9. PASSO 2 — PRÉ-VISUALIZAÇÃO E EDIÇÃO DOS DADOS EXTRAÍDOS
# ============================================================
if st.session_state.dados_extraidos:
    st.markdown("### 👀 3. Confirma os dados antes de gravar no Excel")
    st.caption("Podes corrigir qualquer célula diretamente na tabela abaixo antes de gravar.")

    df_preview = pd.DataFrame(st.session_state.dados_extraidos)
    df_editado = st.data_editor(df_preview, use_container_width=True, num_rows="dynamic")

    if st.button("📥 Preencher e Gerar Excel", use_container_width=True, type="primary"):
        colunas_alvo = [c.strip() for c in campos_usuario.split(",")]
        lista_dados_final = df_editado.to_dict(orient="records")

        try:
            excel_modelo.seek(0)
            excel_bytes = excel_modelo.read()
            output_final, linhas_inseridas = preencher_excel(excel_bytes, lista_dados_final, colunas_alvo)

            st.session_state.excel_pronto = output_final.getvalue()
            st.session_state.linhas_inseridas = linhas_inseridas
            st.session_state.nome_ficheiro_saida = f"Relatorio_Final_{datetime.now().strftime('%d-%m-%Y_%H%M')}.xlsx"

            st.success(f"✨ Sucesso! Adicionámos {linhas_inseridas} nova(s) linha(s) ao documento original.")
        except Exception as ex:
            st.error(f"Erro ao manipular o ficheiro Excel: {ex}")
            logging.error(f"Erro na manipulação de openpyxl: {ex}")

# ============================================================
# 10. PASSO 3 — DOWNLOAD DO FICHEIRO FINAL
# ============================================================
if st.session_state.excel_pronto:
    st.divider()
    st.markdown("### 📥 4. Descarregar Excel Preenchido")
    st.download_button(
        label=f"📥 Descarregar '{st.session_state.nome_ficheiro_saida}'",
        data=st.session_state.excel_pronto,
        file_name=st.session_state.nome_ficheiro_saida,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )
